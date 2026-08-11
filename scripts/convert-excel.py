#!/usr/bin/env python3
"""
Chuyển file Excel "Bài tập điện hạng 3 - Tổng hợp" thành hai file JSON.

Nguyên tắc tách (quan trọng - quyết định việc update app có mất dữ liệu không):

  catalog.json  -> ĐI KÈM CODE. Danh mục 1608 bài: tên bài, chủ đề, độ khó,
                   link denken-ou.com. Cập nhật app là cập nhật file này.
                   Không chứa bất cứ thứ gì riêng của người dùng.

  seed.json     -> DỮ LIỆU KHỞI TẠO. Chỉ dùng đúng một lần ở lần chạy đầu tiên
                   để đổ tiến độ hiện có từ Excel vào kho dữ liệu người dùng.
                   Chứa: trạng thái, ghi chú, link tham khảo, ngày làm, chu kỳ ôn.
                   Từ lần chạy thứ hai trở đi app không bao giờ đọc lại file này.

Hai bên nối nhau bằng `id` sinh từ URL bài tập, nên khi catalog được cập nhật
(thêm bài mới, sửa tên bài) thì tiến độ cũ vẫn khớp đúng vào bài của nó.

Chạy:  python3 scripts/convert-excel.py scripts/source.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# (khoá, tên sheet trong Excel, tên tiếng Nhật, tên tiếng Việt)
SUBJECTS = [
    ("riron", "Riron (理論)", "理論", "Lý thuyết"),
    ("denryoku", "Denryoku (電力)", "電力", "Điện lực"),
    ("kikai", "Kikai (機械)", "機械", "Máy điện"),
    ("houki", "Houki (法規)", "法規", "Pháp quy"),
]

# Cột trong các sheet môn học
COL_NO, COL_TOPIC, COL_TITLE, COL_URL, COL_STATUS = 0, 1, 2, 3, 4
COL_NOTE, COL_DONE_DATE, COL_REF_LINK = 5, 6, 7
COL_SRS_CYCLE, COL_NEXT_REVIEW = 9, 10

STATUS_MAP = {
    "✅ Đúng": "correct",
    "🔄 Sai → Đúng": "relearned",
    "❌ Sai": "wrong",
    "⬜ Chưa làm": "todo",
}

# Chu kỳ ôn tập (ngày) theo từng cấp độ - lấy đúng từ bảng trong Excel
SRS_INTERVALS = [1, 3, 7, 14, 30, 90]


# Giá trị lỗi của Excel: là công thức hỏng chứ không phải nội dung người dùng gõ.
EXCEL_ERROR_RE = re.compile(r"^#(VALUE|REF|DIV/0|N/A|NAME\?|NULL|NUM|SPILL|CALC)!?$")


def clean(value) -> str:
    """Chuẩn hoá một ô Excel thành chuỗi gọn gàng.

    Cố ý KHÔNG dùng NFKC: chuẩn hoá đó sẽ đổi ngoặc full-width （）trong tên
    chủ đề tiếng Nhật thành () nửa chiều, làm sai quy ước hiển thị.
    """
    if value is None:
        return ""
    text = str(value).replace("　", " ").replace("\xa0", " ")
    text = re.sub(r"[ \t\r\n]+", " ", text).strip()
    # Ô #VALUE! không phải ghi chú — bỏ đi để không nhập rác vào app.
    return "" if EXCEL_ERROR_RE.match(text) else text


def clean_freetext(value) -> str:
    """Như clean() nhưng giữ nguyên chữ người dùng gõ.

    Ghi chú là văn bản tự do: gộp khoảng trắng sẽ làm mất xuống dòng và thụt
    đầu dòng trong những ghi chú dài. Chỉ cắt khoảng trắng thừa hai đầu.
    """
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return "" if EXCEL_ERROR_RE.match(text) else text


def to_iso_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def slug_from_url(url: str) -> str:
    """https://denken-ou.com/rironr7-2-1/ -> rironr7-2-1"""
    slug = re.sub(r"^https?://[^/]+/?", "", url).strip("/")
    slug = slug.split("?")[0].split("#")[0]
    return slug.lower()


TITLE_RE = re.compile(
    r"^(?P<stars>[★☆]*)\s*"
    r"(?:《(?P<subject>[^》]*)》)?\s*"
    r"(?:〈(?P<category>[^〉]*)〉)?\s*"
    r"(?:\[(?P<exam>[^\]:]*):(?P<question>[^\]]*)\])?\s*"
    r"(?P<name>.*)$"
)

# Đuôi rác mà trang nguồn gắn vào cuối tên bài
TAIL_RE = re.compile(r"\s*(?:[–—-]\s*Bài tập.*|\|\s*電験王.*)$")


def parse_title(raw: str) -> dict:
    """Tách '★★☆☆☆《理論》〈電磁気〉[R07下:問1]tên bài' thành các phần."""
    match = TITLE_RE.match(raw)
    if not match:
        return {"name": raw, "stars": 0, "category": "", "exam": "", "question": ""}

    name = TAIL_RE.sub("", match.group("name") or "").strip()
    return {
        "name": name or raw,
        "stars": (match.group("stars") or "").count("★"),
        "category": (match.group("category") or "").strip(),
        "exam": (match.group("exam") or "").strip(),
        "question": (match.group("question") or "").strip(),
    }


def cycle_to_level(cycle) -> int:
    """Đổi 'Chu kỳ (ngày)' trong Excel thành cấp độ ôn 0-6."""
    if cycle in (None, ""):
        return 0
    try:
        days = int(float(cycle))
    except (TypeError, ValueError):
        return 0
    if days in SRS_INTERVALS:
        return SRS_INTERVALS.index(days) + 1
    # Giá trị lạ: quy về cấp gần nhất không vượt quá
    level = 0
    for index, interval in enumerate(SRS_INTERVALS):
        if days >= interval:
            level = index + 1
    return level


def read_subject_sheet(worksheet, subject_key: str, seen: Counter) -> list[dict]:
    rows = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        no = clean(row[COL_NO])
        if not no:
            continue  # dòng trống hoặc dòng tiêu đề nhóm chủ đề

        raw_title = clean(row[COL_TITLE])
        url = clean(row[COL_URL])

        slug = slug_from_url(url) if url else f"no-link-{no}"
        base_id = f"{subject_key}:{slug}"
        seen[base_id] += 1
        # 24 bài liên môn được liệt kê ở 2 chủ đề khác nhau -> giữ cả hai,
        # phân biệt bằng hậu tố để mỗi dòng có tiến độ riêng như trong Excel.
        item_id = base_id if seen[base_id] == 1 else f"{base_id}~{seen[base_id]}"

        rows.append(
            {
                "id": item_id,
                "subject": subject_key,
                "topic": clean(row[COL_TOPIC]),
                "no": int(float(no)) if re.fullmatch(r"[\d.]+", no) else 0,
                "url": url,
                "raw_title": raw_title,
                **parse_title(raw_title),
                # phần dưới đây là dữ liệu người dùng, sẽ tách sang seed.json
                "_status": STATUS_MAP.get(clean(row[COL_STATUS]), "todo"),
                "_note": clean_freetext(row[COL_NOTE]),
                "_ref_link": clean(row[COL_REF_LINK]),
                "_done_date": to_iso_date(row[COL_DONE_DATE]),
                "_srs_level": cycle_to_level(row[COL_SRS_CYCLE]),
                "_next_review": to_iso_date(row[COL_NEXT_REVIEW]),
            }
        )
    return rows


def read_quizlet_sheet(worksheet) -> tuple[list[dict], list[dict]]:
    """Sheet Quizlet gồm 3 khối: bộ thẻ, bảng kanji, bảng từ vựng."""
    decks, vocab = [], []
    section = "decks"

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        cells = [clean(c) for c in row]
        if not any(cells):
            continue

        # Nhận diện dòng tiêu đề để biết đang sang khối mới
        if cells[0] in ("Kanji", "Từ vựng / Kanji"):
            section = "vocab"
            continue
        if cells[0] == "Mon":
            section = "decks"
            continue

        if section == "decks":
            if not cells[2].startswith("http"):
                continue
            decks.append(
                {
                    "id": f"deck-{len(decks) + 1}",
                    "subject": cells[0],
                    "name": cells[1],
                    "url": cells[2],
                    "remaining": int(float(cells[3])) if cells[3] else 0,
                    "total": int(float(cells[4])) if cells[4] else 0,
                }
            )
        else:
            # Cột 0 đôi khi là '過電流 (かでんりゅう)' -> tách phần trong ngoặc
            term, reading = cells[0], cells[1]
            bracket = re.match(r"^(.*?)\s*[（(]([^）)]*)[）)]\s*$", term)
            if bracket:
                term, reading = bracket.group(1), bracket.group(2)
            vocab.append(
                {
                    "id": f"vocab-{len(vocab) + 1}",
                    "term": term,
                    "reading": reading,
                    "meaning": cells[2],
                    "hint": cells[3],
                }
            )
    return decks, vocab


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "scripts" / "source.xlsx"
    if not source.exists():
        sys.exit(f"Không tìm thấy file Excel: {source}")

    workbook = openpyxl.load_workbook(source, data_only=True)

    catalog_items: list[dict] = []
    progress: dict[str, dict] = {}
    seen: Counter = Counter()

    for key, sheet_name, ja_name, vi_name in SUBJECTS:
        for row in read_subject_sheet(workbook[sheet_name], key, seen):
            catalog_items.append(
                {
                    "id": row["id"],
                    "subject": row["subject"],
                    "topic": row["topic"],
                    "no": row["no"],
                    "name": row["name"],
                    "stars": row["stars"],
                    "category": row["category"],
                    "exam": row["exam"],
                    "question": row["question"],
                    "url": row["url"],
                }
            )

            # Chỉ ghi vào seed những bài thực sự có tiến độ, cho file gọn
            entry = {
                k: v
                for k, v in {
                    "status": row["_status"],
                    "note": row["_note"],
                    "refLink": row["_ref_link"],
                    "doneDate": row["_done_date"],
                    "srsLevel": row["_srs_level"],
                    "nextReview": row["_next_review"],
                }.items()
                if v not in (None, "", 0, "todo")
            }
            if entry:
                progress[row["id"]] = entry

    decks, vocab = read_quizlet_sheet(workbook["Quizlet"])

    catalog = {
        "version": 1,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": source.name,
        "srsIntervals": SRS_INTERVALS,
        "subjects": [
            {"key": k, "name": ja, "viName": vi, "sheet": sheet}
            for k, sheet, ja, vi in SUBJECTS
        ],
        "items": catalog_items,
    }

    seed = {
        "version": 1,
        "generatedAt": catalog["generatedAt"],
        "progress": progress,
        "decks": decks,
        "vocab": vocab,
    }

    catalog_path = ROOT / "src" / "data" / "catalog.json"
    seed_path = ROOT / "src" / "data" / "seed.json"
    catalog_path.parent.mkdir(parents=True, exist_ok=True)

    catalog_path.write_text(
        json.dumps(catalog, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    seed_path.write_text(
        json.dumps(seed, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    # Báo cáo để đối chiếu với Excel gốc
    by_subject = Counter(item["subject"] for item in catalog_items)
    by_status = Counter(
        progress.get(item["id"], {}).get("status", "todo") for item in catalog_items
    )
    print(f"catalog.json : {len(catalog_items)} bài  ({catalog_path})")
    for key, _, ja, vi in SUBJECTS:
        print(f"    {ja} {vi:<10} {by_subject[key]:>5}")
    print(f"seed.json    : {len(progress)} bài có tiến độ, "
          f"{len(decks)} bộ Quizlet, {len(vocab)} từ vựng")
    print(f"    trạng thái: {dict(by_status)}")
    print(f"    ghi chú: {sum(1 for e in progress.values() if e.get('note'))}"
          f" | link tham khảo: {sum(1 for e in progress.values() if e.get('refLink'))}"
          f" | đang trong chu kỳ ôn: {sum(1 for e in progress.values() if e.get('nextReview'))}")


if __name__ == "__main__":
    main()
