#!/usr/bin/env python3
"""
Dựng src/data/answers.json — bảng đáp án cho chức năng thi thử.

Nhận vào một trong hai nguồn:

  CSV    file có cột `dap_an_1` (và `dap_an_2` cho B問題), giá trị 1..5
  Excel  file Excel có thêm cột "Đáp án (a)" và "Đáp án (b)"

A問題 chỉ có một ý -> chỉ điền `dap_an_1`.
B問題 có hai ý (a) và (b) -> điền cả `dap_an_1` và `dap_an_2`.
Đáp án của mỗi câu vì thế là một mảng 1 hoặc 2 phần tử.

Chạy:
    python3 scripts/build-answers.py scripts/mau-dap-an.csv
    python3 scripts/build-answers.py scripts/source.xlsx

Đáp án là dữ liệu chung cho mọi người dùng — giống link bài tập — nên nó nằm
trong `src/data/` đi kèm code, không phải trong thư mục dữ liệu cá nhân.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "src" / "data" / "catalog.json"
TARGET = ROOT / "src" / "data" / "answers.json"

# Hai cột đáp án nếu bạn thêm vào file Excel (đánh số từ 0)
COL_ANSWER_A = 12   # ý (a), hoặc đáp án duy nhất của A問題
COL_ANSWER_B = 13   # ý (b), chỉ B問題 mới có


def valid_ids() -> set[str]:
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    return {item["id"] for item in catalog["items"]}


def parse_choice(raw) -> int | None:
    """Chấp nhận '3', '3.0', '(3)', '③' — trả về 1..5, sai thì None."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    circled = "①②③④⑤"
    if text in circled:
        return circled.index(text) + 1

    match = re.search(r"[1-5]", text)
    return int(match.group()) if match else None


def collect(first, second, item_id: str, allowed: set[str],
            answers: dict[str, list[int]], problems: Counter) -> None:
    """Ghép hai ý thành mảng đáp án của một câu."""
    a = parse_choice(first)
    b = parse_choice(second)
    if a is None and b is None:
        return
    if item_id not in allowed:
        problems["id không có trong danh mục"] += 1
        return
    if a is None:
        # Có ý (b) mà thiếu ý (a) thì câu này chấm được một nửa, vẫn giữ lại
        # nhưng phải báo để người điền biết mà bổ sung.
        problems["thiếu ý (a) dù đã có ý (b)"] += 1
        return
    answers[item_id] = [a] if b is None else [a, b]


def from_csv(path: Path, allowed: set[str]) -> tuple[dict[str, list[int]], Counter]:
    answers: dict[str, list[int]] = {}
    problems: Counter = Counter()

    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            collect(
                row.get("dap_an_1") or row.get("dap_an"),
                row.get("dap_an_2"),
                (row.get("id") or "").strip(),
                allowed,
                answers,
                problems,
            )
    return answers, problems


def from_excel(path: Path, allowed: set[str]) -> tuple[dict[str, list[int]], Counter]:
    import openpyxl

    # Sinh id y hệt cách convert-excel.py làm: lấy từ link bài, thêm hậu tố ~2
    # cho những bài liên môn xuất hiện hai lần.
    subjects = [
        ("riron", "Riron (理論)"),
        ("denryoku", "Denryoku (電力)"),
        ("kikai", "Kikai (機械)"),
        ("houki", "Houki (法規)"),
    ]

    workbook = openpyxl.load_workbook(path, data_only=True)
    answers: dict[str, list[int]] = {}
    problems: Counter = Counter()

    for key, sheet_name in subjects:
        if sheet_name not in workbook.sheetnames:
            continue
        seen: Counter = Counter()
        for row in workbook[sheet_name].iter_rows(min_row=3, values_only=True):
            no = str(row[0] or "").strip()
            if not no:
                continue

            url = str(row[3] or "").strip()
            slug = re.sub(r"^https?://[^/]+/?", "", url).strip("/").lower() or f"no-link-{no}"
            base = f"{key}:{slug}"
            seen[base] += 1
            item_id = base if seen[base] == 1 else f"{base}~{seen[base]}"

            collect(
                row[COL_ANSWER_A] if len(row) > COL_ANSWER_A else None,
                row[COL_ANSWER_B] if len(row) > COL_ANSWER_B else None,
                item_id,
                allowed,
                answers,
                problems,
            )

    return answers, problems


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    source = Path(sys.argv[1])
    if not source.exists():
        sys.exit(f"Không tìm thấy file: {source}")

    allowed = valid_ids()
    if source.suffix.lower() in (".xlsx", ".xlsm"):
        answers, problems = from_excel(source, allowed)
    else:
        answers, problems = from_csv(source, allowed)

    TARGET.write_text(
        json.dumps(
            {
                "version": 1,
                "note": (
                    "Bảng đáp án cho chức năng thi thử. Khoá là id bài (giống "
                    "catalog.json), giá trị là MẢNG đáp án theo từng ý: A問題 "
                    "một phần tử, B問題 hai phần tử cho ý (a) và (b). "
                    "Câu nào chưa có ở đây thì bài thi bỏ qua không chấm, và điểm "
                    "được quy về thang 100 trên phần chấm được. "
                    f"Sinh từ {source.name} bằng scripts/build-answers.py."
                ),
                "answers": dict(sorted(answers.items())),
            },
            ensure_ascii=False,
            indent=1,
        ),
        encoding="utf-8",
    )

    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    with_exam = [i for i in catalog["items"] if i["exam"]]
    per_exam: Counter = Counter()
    for item in with_exam:
        if item["id"] in answers:
            per_exam[item["exam"]] += 1

    subs = sum(len(v) for v in answers.values())
    print(f"answers.json : {len(answers)} câu ({subs} ý) có đáp án"
          f" trên tổng {len(with_exam)} câu")
    two = sum(1 for v in answers.values() if len(v) == 2)
    print(f"   trong đó {two} câu có đủ 2 ý (B問題)")
    if problems:
        print("   bỏ qua:", dict(problems))
    if per_exam:
        print("   kỳ thi đã có đáp án:")
        for exam, count in sorted(per_exam.items(), key=lambda kv: -kv[1])[:10]:
            total = sum(1 for i in with_exam if i["exam"] == exam)
            print(f"      {exam:<8} {count:>3}/{total}")


if __name__ == "__main__":
    main()
