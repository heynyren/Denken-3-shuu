#!/usr/bin/env python3
"""Nạp bản dịch tiếng Việt từ file Excel vào danh mục đang có.

    python3 scripts/nap-tieng-viet.py scripts/source.xlsx --thu
    python3 scripts/nap-tieng-viet.py scripts/source.xlsx

Bản dịch nằm ngay trong ô "Tên bài tập" của Excel, sau dấu gạch ngang:

    ★★☆☆☆《理論》〈電磁気〉[R07下:問1]…に関する計算問題 – Bài tập tính toán về…

Vì sao cần script riêng thay vì chạy lại convert-excel.py: danh mục hiện tại đã
qua `sua-danh-muc.py` vá lại kỳ thi, số câu và ba đường link lệch. Chạy lại
convert từ đầu là mất hết những chỗ vá đó. Script này chỉ **thêm** trường
`nameVi`, không đụng gì khác.

Ghép theo link vì link là thứ ổn định nhất; link nào không khớp thì ghép theo
tên bài tiếng Nhật.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "src" / "data" / "catalog.json"

SHEETS = {
    "Riron (理論)": "riron",
    "Denryoku (電力)": "denryoku",
    "Kikai (機械)": "kikai",
    "Houki (法規)": "houki",
}
COL_TITLE, COL_URL = 2, 3

# Bản dịch tiếng Việt nằm cuối ô tên bài, người dùng viết theo hai kiểu:
#   …に関する計算問題 – Bài tập tính toán về…      (dấu gạch ngang)
#   …に関する計算問題 (Bài tính toán về…)          (ngoặc đơn, hay gặp ở 電熱)
# Đòi phải có chữ tiếng Việt ở trong để không cắt nhầm dấu gạch ngang hay ngoặc
# của chính tên tiếng Nhật (ví dụ "D-A 変換器", hay ngoặc toàn角 "（静電界）").
TRANSLATION_RE = re.compile(
    r"\s(?:[–—]\s*((?=[^\n]*[À-ỹ])[^\n]+)|\((?=[^)]*[À-ỹ])([^)]+)\))\s*$"
)
# Phần đánh dấu ở đầu tên: sao, môn, nhóm, [kỳ thi:câu]
HEAD_RE = re.compile(r"^[★☆]*\s*(?:《[^》]*》)?\s*(?:〈[^〉]*〉)?\s*(?:\[[^\]]*\])?\s*")


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry_run = "--thu" in sys.argv
    if not args:
        sys.exit(__doc__)

    book = openpyxl.load_workbook(args[0], read_only=True, data_only=True)

    by_url: dict[str, str] = {}
    by_name: dict[str, str] = {}
    total = 0

    for sheet in SHEETS:
        if sheet not in book.sheetnames:
            continue
        for row in book[sheet].iter_rows(min_row=3, values_only=True):
            title = str(row[COL_TITLE] or "").strip()
            if not title:
                continue
            total += 1

            found = TRANSLATION_RE.search(title)
            if not found:
                continue
            vietnamese = (found.group(1) or found.group(2)).strip()
            japanese = HEAD_RE.sub("", title[: found.start()]).strip()

            url = str(row[COL_URL] or "").strip()
            if url:
                by_url[url] = vietnamese
            if japanese:
                by_name.setdefault(japanese, vietnamese)

    print(f"Đọc {total} dòng, có {len(by_url)} bản dịch kèm link.")

    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    hit_url = hit_name = miss = 0
    missing: list[str] = []

    for item in catalog["items"]:
        vietnamese = by_url.get(item["url"]) or by_name.get(item["name"])
        if vietnamese:
            item["nameVi"] = vietnamese
            if by_url.get(item["url"]):
                hit_url += 1
            else:
                hit_name += 1
        else:
            item.setdefault("nameVi", "")
            miss += 1
            if len(missing) < 25:
                missing.append(f"{item['exam']} {item['subject']} {item['question']}")

    print(f"Ghép theo link: {hit_url} · ghép theo tên: {hit_name} · chưa có: {miss}")
    if missing:
        print("Bài chưa có bản dịch:")
        for line in missing:
            print("   " + line)

    if dry_run:
        print("\n(--thu: không ghi file)")
        return

    CATALOG.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )
    print(f"\nĐã ghi {CATALOG.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
